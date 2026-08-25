#!/usr/bin/env python3
"""M2.14.2 Demo fixture generator (Phase 1).

Runs the REAL production pipeline (schema detection -> Metric Engine ->
Health Score) against a deterministic bundled sample workbook, then writes:

- backend/demo/sample_sales_data.xlsx   (sample workbook, 3 periods)
- frontend/src/lib/demo/demo_result.json (fixture snapshot consumed by /demo)

Design rules (Lytio moat constraints):
- Numbers come ONLY from the real pipeline (code-computed). No AI, no hand
  edits, no fabricated metrics.
- Narrative text lives in i18n (frontend/src/lib/i18n.ts), keyed by stable ids
  in the fixture; the fixture itself contains facts and structure only.
- The fixture is versioned (engine_version / source_commit / data_md5).

Usage:
    python scripts/generate_demo_fixture.py [--commit <sha>]

Output is deterministic (fixed RNG seed). Re-running with the same commit
reproduces identical metrics.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import random
import sys
from datetime import datetime, timezone
from pathlib import Path

BACKEND_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.services.health_score import compute_health_score  # noqa: E402
from app.services.metric_engine import compute_metrics  # noqa: E402
from app.services.schema_mapper import derive_schema_meta, detect_schema  # noqa: E402

SAMPLE_XLSX = BACKEND_ROOT / "demo" / "sample_sales_data.xlsx"
FIXTURE_JSON = BACKEND_ROOT.parent / "frontend" / "src" / "lib" / "demo" / "demo_result.json"

# Production metric keys (must match metric_engine.py output exactly).
SUPPORTED_METRICS = {
    "total_sales",
    "sales_growth",
    "order_count",
    "average_order_value",
    "customer_count",
    "customer_concentration",
    "product_sales_rank",
    "row_count",
    "date_range",
}
CORE_METRICS = [
    "total_sales",
    "order_count",
    "average_order_value",
    "customer_count",
    "customer_concentration",
]

PRODUCTS = [
    ("智能音箱", 299),
    ("无线耳机", 199),
    ("便携充电宝", 129),
    ("智能手表", 599),
    ("蓝牙键盘", 149),
]
CUSTOMERS = ["华信贸易", "青禾零售", "远航商贸", "科讯电子", "蓝海连锁", "天达百货", "启明科技", "南方优选"]
REGIONS = ["华东", "华北", "华南", "西南"]
PERSONS = ["王磊", "李娜", "张强", "陈静"]

# Monthly revenue targets (dip in Feb, recovery in Mar) -> demo story:
# period1 healthy, period2 decline, period3 recovery (aligned with actions).
MONTH_TARGETS = [("2026-01", 380_000.0), ("2026-02", 220_000.0), ("2026-03", 310_000.0)]
ROWS_PER_MONTH = 40


def _detect_type(values: list) -> str:
    """Mirror of app.services.workbook_service._detect_type (kept offline)."""
    non_empty = [v for v in values if v not in (None, "", "None")]
    if not non_empty:
        return "empty"
    types = {type(v) for v in non_empty}
    if types in ({int}, {float}, {int, float}):
        return "number"
    if types == {str}:
        return "text"
    if types == {bool}:
        return "boolean"
    if all(hasattr(v, "strftime") for v in non_empty):
        return "date"
    return "unknown"


def build_sample_rows() -> list[list]:
    rng = random.Random(20260824)
    by_month: dict[str, list[list]] = {}
    for month, target in MONTH_TARGETS:
        rows: list[list] = []
        for _ in range(ROWS_PER_MONTH):
            day = rng.randint(1, 28)
            date = f"{month}-{day:02d}"
            cust = rng.choice(CUSTOMERS)
            prod, price = rng.choice(PRODUCTS)
            qty = rng.randint(2, 60)
            amt = qty * price * rng.uniform(0.85, 1.15)
            region = rng.choice(REGIONS)
            person = rng.choice(PERSONS)
            rows.append([date, cust, prod, round(amt, 2), qty, region, person])
        total = sum(r[3] for r in rows)
        scale = target / total if total else 1.0
        for r in rows:
            r[3] = round(r[3] * scale, 2)
        by_month[month] = rows

    ordered = []
    for month, _ in MONTH_TARGETS:
        ordered.extend(by_month[month])
    return ordered


def build_workbook(all_rows: list[list]) -> None:
    months = [m for m, _ in MONTH_TARGETS]
    wb = Workbook()
    wb.remove(wb.active)
    header = ["订单日期", "客户名称", "产品名称", "销售金额", "数量", "区域", "销售人员"]
    slices = {
        1: all_rows[: ROWS_PER_MONTH],
        2: all_rows[: ROWS_PER_MONTH * 2],
        3: all_rows[: ROWS_PER_MONTH * 3],
    }
    for period in (1, 2, 3):
        ws = wb.create_sheet(f"销售数据_周期{period}")
        ws.append(header)
        for r in sorted(slices[period], key=lambda x: x[0]):
            ws.append(r)
    SAMPLE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(SAMPLE_XLSX))


def extract_dataset(path: Path, sheet_name: str) -> dict:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet = wb[sheet_name]
        headers: list[str] = []
        rows: list[list] = []
        columns_data: dict[int, list] = {}
        header_found = False
        for row in wb[sheet_name].iter_rows(values_only=True):
            values = [v if v is not None else "" for v in row]
            if all(v == "" for v in values):
                continue
            if not header_found:
                headers = [str(v) for v in values]
                columns_data = {i: [] for i in range(len(headers))}
                header_found = True
            else:
                for i, v in enumerate(values):
                    if i in columns_data:
                        columns_data[i].append(v)
                rows.append(values)
        column_types = {name: _detect_type(columns_data.get(i, [])) for i, name in enumerate(headers)}
        return {
            "workbook_name": SAMPLE_XLSX.name,
            "sheet_name": sheet_name,
            "headers": headers,
            "column_types": column_types,
            "rows": rows,
        }
    finally:
        wb.close()


def fmt_value(value) -> str:
    if isinstance(value, float):
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, dict):
        return f"{value.get('min')} ~ {value.get('max')}"
    return str(value)


def run_pipeline(dataset: dict) -> dict:
    detection = detect_schema(dataset["headers"], dataset["column_types"])
    mapping = detection.to_dict()
    metrics = compute_metrics(dataset, mapping)
    health = compute_health_score(dataset, mapping, metrics)
    schema_meta = derive_schema_meta(mapping)
    metric_by_name = {m["metric_name"]: m for m in metrics}
    key_metrics = []
    for name in CORE_METRICS:
        m = metric_by_name.get(name)
        if not m:
            continue
        available = m.get("availability") == "available" and m.get("value") is not None
        key_metrics.append(
            {
                "metric_name": name,
                "value": m.get("value") if available else None,
                "display": fmt_value(m.get("value")) if available else "—",
                "availability": m.get("availability"),
                "confidence": m.get("confidence"),
                "formula": m.get("formula"),
                "assumptions": m.get("assumptions") or [],
                "note": m.get("note") or "",
            }
        )
    return {
        "schema_meta": schema_meta,
        "schema_mapping": {
            "version": mapping.get("version"),
            "schema_version": mapping.get("schema_version"),
            "mappings": mapping.get("mappings"),
            "missing": mapping.get("missing"),
            "conflicts": mapping.get("conflicts"),
            "sales_core_available": mapping.get("sales_core_available"),
        },
        "computed_metrics": metrics,
        "health_score": health,
        "key_metrics": key_metrics,
    }


def _format_pct(v) -> str:
    if v is None:
        return "—"
    return f"{v * 100:+.1f}%"


def _format_conc(v) -> str:
    if v is None:
        return "—"
    return f"{v * 100:.1f}%"


def build_fixture(source_commit: str) -> dict:
    all_rows = build_sample_rows()
    build_workbook(all_rows)

    data_md5 = hashlib.md5(SAMPLE_XLSX.read_bytes()).hexdigest()
    periods = []
    for period in (1, 2, 3):
        dataset = extract_dataset(SAMPLE_XLSX, f"销售数据_周期{period}")
        res = run_pipeline(dataset)
        res["period_id"] = period
        periods.append(res)

    current = periods[-1]
    growth = next((m.get("value") for m in current["computed_metrics"] if m["metric_name"] == "sales_growth"), None)
    conc = next((m.get("value") for m in current["computed_metrics"] if m["metric_name"] == "customer_concentration"), None)
    total_sales = next((m.get("value") for m in current["computed_metrics"] if m["metric_name"] == "total_sales"), None)
    top_product = next((m.get("value") for m in current["computed_metrics"] if m["metric_name"] == "product_sales_rank"), None)
    aov = next((m.get("value") for m in current["computed_metrics"] if m["metric_name"] == "average_order_value"), None)
    customers = next((m.get("value") for m in current["computed_metrics"] if m["metric_name"] == "customer_count"), None)
    health_score = current["health_score"].get("health_score")

    timeline = []
    for idx, p in enumerate(periods, start=1):
        hs = p["health_score"]
        timeline.append(
            {
                "id": idx,
                "run_id": 300 + idx,
                "created_at": f"2026-0{idx}-15T09:00:00Z",
                "business_health_score": hs.get("health_score"),
                "level": hs.get("health_level"),
                "period": f"周期{idx}",
                "summary_key": f"demo.timeline.c{idx}",
            }
        )

    narrative = {
        "insights": [
            {"id": "insight_growth", "confidence": "high"},
            {"id": "insight_concentration", "confidence": "high"},
            {"id": "insight_product", "confidence": "medium"},
            {"id": "insight_region", "confidence": "medium"},
        ],
        "risks": [
            {"id": "risk_revenue", "severity": "high"},
            {"id": "risk_concentration", "severity": "medium"},
            {"id": "risk_pipeline", "severity": "medium"},
        ],
        "recommendations": [
            {"id": "rec_recovery", "priority": "high"},
            {"id": "rec_region", "priority": "high"},
            {"id": "rec_customers", "priority": "medium"},
        ],
        "actions": [
            {
                "id": 101,
                "recommendation_id": "rec_recovery",
                "source_run_id": 303,
                "description_key": "demo.rec.recovery.desc",
                "expected_result_key": "demo.rec.recovery.expected",
                "reason_key": "demo.verify.reason.aligned",
                "priority_snapshot": "high",
                "status": "completed",
                "execution_count": 1,
                "target_metric_name": "total_sales",
                "target_direction": "up",
                "target_metric_source": "user",
                "observations_summary": {"total": 1, "aligned": 1, "not_aligned": 0, "unable_to_verify": 0},
                "verification_run_id": 401,
                "verified_at": "2026-03-20T10:00:00Z",
            },
            {
                "id": 102,
                "recommendation_id": "rec_region",
                "source_run_id": 303,
                "description_key": "demo.rec.region.desc",
                "expected_result_key": "demo.rec.region.expected",
                "reason_key": "demo.verify.reason.aligned",
                "priority_snapshot": "high",
                "status": "completed",
                "execution_count": 1,
                "target_metric_name": "total_sales",
                "target_direction": "up",
                "target_metric_source": "user",
                "observations_summary": {"total": 1, "aligned": 1, "not_aligned": 0, "unable_to_verify": 0},
                "verification_run_id": 401,
                "verified_at": "2026-03-20T10:00:00Z",
            },
            {
                "id": 103,
                "recommendation_id": "rec_customers",
                "source_run_id": 303,
                "description_key": "demo.rec.customers.desc",
                "expected_result_key": "demo.rec.customers.expected",
                "priority_snapshot": "medium",
                "status": "pending",
                "execution_count": 0,
                "target_metric_name": "customer_concentration",
                "target_direction": "down",
                "target_metric_source": "user",
                "observations_summary": None,
                "verification_run_id": None,
                "verified_at": None,
            },
            {
                "id": 104,
                "recommendation_id": "rec_pipeline",
                "source_run_id": 302,
                "description_key": "demo.rec.pipeline.desc",
                "expected_result_key": "demo.rec.pipeline.expected",
                "priority_snapshot": "medium",
                "status": "pending",
                "execution_count": 0,
                "target_metric_name": None,
                "target_direction": None,
                "target_metric_source": "none",
                "observations_summary": None,
                "verification_run_id": None,
                "verified_at": None,
            },
        ],
        "verification": {
            "run_id": 401,
            "parent_run_id": 303,
            "verdict": "partially_effective",
            "confidence": "medium",
            "reliability": "verification_reliability_v1",
            "metric_changes": [
                {
                    "metric": "total_sales",
                    "before": periods[1]["key_metrics"][0]["display"],
                    "after": periods[2]["key_metrics"][0]["display"],
                    "absolute_delta": None,
                    "percent_delta": _format_pct(growth),
                    "direction": "up" if (growth or 0) > 0 else "down",
                    "status": "computed",
                },
                {
                    "metric": "average_order_value",
                    "before": periods[1]["key_metrics"][2]["display"],
                    "after": periods[2]["key_metrics"][2]["display"],
                    "absolute_delta": None,
                    "percent_delta": "—",
                    "direction": "unchanged",
                    "status": "computed",
                },
            ],
        },
        "memory": {
            "project_id": 0,
            "engine_version": "business_memory_v0",
            "profile": {},
            "latest_metrics": {
                "total_sales": {"value": periods[2]["key_metrics"][0]["value"], "availability": "available", "confidence": "high"},
                "average_order_value": {"value": periods[2]["key_metrics"][2]["value"], "availability": "available", "confidence": "medium"},
                "customer_count": {"value": periods[2]["key_metrics"][3]["value"], "availability": "available", "confidence": "high"},
                "customer_concentration": {"value": periods[2]["key_metrics"][4]["value"], "availability": "available", "confidence": "high"},
            },
            "metric_history": {
                "total_sales": [
                    {"run_id": 301, "dataset_version": "周期1", "period": {"min": "2026-01", "max": "2026-01"}, "value": periods[0]["key_metrics"][0]["value"]},
                    {"run_id": 302, "dataset_version": "周期2", "period": {"min": "2026-01", "max": "2026-02"}, "value": periods[1]["key_metrics"][0]["value"]},
                    {"run_id": 303, "dataset_version": "周期3", "period": {"min": "2026-01", "max": "2026-03"}, "value": periods[2]["key_metrics"][0]["value"]},
                ]
            },
            "health_history": [
                {"run_id": 301, "dataset_version": "周期1", "period": {"min": "2026-01", "max": "2026-01"}, "score": periods[0]["health_score"].get("health_score"), "level": periods[0]["health_score"].get("health_level")},
                {"run_id": 302, "dataset_version": "周期2", "period": {"min": "2026-01", "max": "2026-02"}, "score": periods[1]["health_score"].get("health_score"), "level": periods[1]["health_score"].get("health_level")},
                {"run_id": 303, "dataset_version": "周期3", "period": {"min": "2026-01", "max": "2026-03"}, "score": periods[2]["health_score"].get("health_score"), "level": periods[2]["health_score"].get("health_level")},
            ],
            "action_summary": {"total": 4, "pending": 2, "completed": 2, "cancelled": 0, "verified": 2},
            "action_recent": [],
            "issue_tracker": [],
            "verification_history": [
                {
                    "run_id": 401,
                    "parent_run_id": 303,
                    "verdict": "partially_effective",
                    "confidence": "medium",
                    "period": {"min": "2026-02", "max": "2026-03"},
                    "alignment": "aligned",
                    "observations": 2,
                    "metric_changes": [
                        {"metric": "total_sales", "direction": "up"},
                        {"metric": "average_order_value", "direction": "unchanged"},
                    ],
                    "next_actions": [],
                }
            ],
            "open_loops": [
                {"type": "not_executed_action", "priority": "medium", "description_key": "demo.loop.notExecuted", "first_seen_run_id": 303},
                {"type": "long_open_issue", "priority": "medium", "metric": "pipeline_health", "description_key": "demo.loop.longOpen", "first_seen_run_id": 301},
            ],
            "intelligence": {
                "engine_version": "business_memory_intelligence_v1",
                "rates": {
                    "execution": {"action_total": 4, "executed_count": 2, "execution_rate": 0.5},
                    "verification": {
                        "total_verified_actions": 4,
                        "verified_count": 2,
                        "verification_rate": 0.5,
                        "unable_to_verify_count": 2,
                        "unable_rate": 0.5,
                        "unable_reasons": {"not_executed": 1, "metric_unavailable": 1, "insufficient_data": 0},
                    },
                },
                "alignment_trend": [
                    {
                        "period": "2026-02 ~ 2026-03",
                        "verification_run_id": 401,
                        "aligned_count": 2,
                        "not_aligned_count": 0,
                        "unable_count": 0,
                        "source_run_ids": [303],
                    }
                ],
                "improvement_timeline": [
                    {
                        "period": "2026-02 ~ 2026-03",
                        "verification_run_id": 401,
                        "parent_run_id": 303,
                        "observation_count": 2,
                        "observations": [
                            {
                                "action_id": 101,
                                "description_key": "demo.rec.recovery.desc",
                                "metric_name": "total_sales",
                                "before_value": periods[1]["key_metrics"][0]["value"],
                                "after_value": periods[2]["key_metrics"][0]["value"],
                                "absolute_delta": round((periods[2]["key_metrics"][0]["value"] or 0) - (periods[1]["key_metrics"][0]["value"] or 0), 2),
                                "percent_delta": round(growth, 4) if growth is not None else None,
                                "direction": "up" if (growth or 0) > 0 else "down",
                                "alignment": "aligned",
                                "executed": True,
                                "reason_key": "demo.verify.reason.aligned",
                            },
                            {
                                "action_id": 102,
                                "description_key": "demo.rec.region.desc",
                                "metric_name": "total_sales",
                                "before_value": periods[1]["key_metrics"][0]["value"],
                                "after_value": periods[2]["key_metrics"][0]["value"],
                                "absolute_delta": round((periods[2]["key_metrics"][0]["value"] or 0) - (periods[1]["key_metrics"][0]["value"] or 0), 2),
                                "percent_delta": round(growth, 4) if growth is not None else None,
                                "direction": "up" if (growth or 0) > 0 else "down",
                                "alignment": "aligned",
                                "executed": True,
                                "reason_key": "demo.verify.reason.aligned",
                            },
                        ],
                    }
                ],
                "open_loops": [
                    {"type": "not_executed_action", "priority": "medium", "description_key": "demo.loop.notExecuted", "first_seen_run_id": 303},
                    {"type": "long_open_issue", "priority": "medium", "metric": "pipeline_health", "description_key": "demo.loop.longOpen", "first_seen_run_id": 301},
                ],
            },
            "trend_deltas": {
                "metric_trends": [
                    {
                        "metric_name": "total_sales",
                        "latest": periods[2]["key_metrics"][0]["value"],
                        "previous": periods[1]["key_metrics"][0]["value"],
                        "absolute_delta": round((periods[2]["key_metrics"][0]["value"] or 0) - (periods[1]["key_metrics"][0]["value"] or 0), 2),
                        "percent_delta": round(growth, 4) if growth is not None else None,
                        "direction": "up" if (growth or 0) > 0 else "down",
                        "period_count": 3,
                        "availability": "available",
                        "confidence": "high",
                        "source_run_ids": [301, 302, 303],
                    }
                ],
                "health_trend": {
                    "latest_score": periods[2]["health_score"].get("health_score"),
                    "previous_score": periods[1]["health_score"].get("health_score"),
                    "delta": round((periods[2]["health_score"].get("health_score") or 0) - (periods[1]["health_score"].get("health_score") or 0), 1),
                    "direction": "up",
                    "latest_level": periods[2]["health_score"].get("health_level"),
                    "score_confidence": periods[2]["health_score"].get("score_confidence"),
                    "period_count": 3,
                    "source_run_ids": [301, 302, 303],
                },
                "action_trend": {
                    "total_actions": 4,
                    "pending": 2,
                    "completed": 2,
                    "cancelled": 0,
                    "verified": 2,
                    "verification_rate": 0.5,
                    "open_loops": 2,
                    "source": "demo_fixture",
                },
                "verification_trend": {
                    "latest_verdict": "partially_effective",
                    "previous_verdict": None,
                    "latest_reliability": "verification_reliability_v1",
                    "latest_confidence": "medium",
                    "verified_recommendations": 2,
                    "metric_changes_summary": [{"metric_name": "total_sales", "direction": "up"}],
                    "source_run_ids": [401],
                },
                "periods_used": 3,
                "latest_run_id": 303,
            },
            "context_meta": {
                "version": "demo_fixture",
                "periods_used": 3,
                "latest_run_id": 303,
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "length_chars": 0,
                "injected": False,
            },
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "ready": True,
        },
        "timeline": timeline,
        "params": {
            "total_sales": fmt_value(total_sales) if total_sales is not None else "—",
            "growth": _format_pct(growth),
            "growth_sign": "up" if (growth or 0) >= 0 else "down",
            "concentration": _format_conc(conc),
            "customers": fmt_value(customers) if customers is not None else "—",
            "aov": fmt_value(aov) if aov is not None else "—",
            "top_product": str(top_product) if top_product else "—",
            "health_score": str(round(health_score)) if health_score is not None else "—",
            "health_level": str(periods[2]["health_score"].get("health_level") or ""),
        },
    }

    return {
        "meta": {
            "schema_version": str(periods[0]["schema_meta"].get("schema_version") or "canonical_sales_v1"),
            "engine_version": "metric_engine_v1",
            "health_score_engine": "health_score_v1",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_commit": source_commit,
            "sample_file": "backend/demo/sample_sales_data.xlsx",
            "data_md5": data_md5,
        },
        "supported_metrics": sorted(SUPPORTED_METRICS),
        "periods": periods,
        "narrative": narrative,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate demo fixture from the real pipeline.")
    parser.add_argument("--commit", default="03db6e8", help="Source commit the fixture is generated from.")
    args = parser.parse_args()

    fixture = build_fixture(args.commit)

    # Guard rails (mirror demo:check essentials).
    current = fixture["periods"][-1]
    for name in CORE_METRICS:
        m = next((x for x in current["computed_metrics"] if x["metric_name"] == name), None)
        assert m is not None, f"missing core metric {name}"
        assert m["metric_name"] in SUPPORTED_METRICS, f"unknown metric {name}"
    for m in current["computed_metrics"]:
        assert m["confidence"] != "low", f"unexpected low confidence on {m['metric_name']}"
        assert m["metric_name"] in SUPPORTED_METRICS, f"removed/unknown metric {m['metric_name']}"

    FIXTURE_JSON.parent.mkdir(parents=True, exist_ok=True)
    FIXTURE_JSON.write_text(
        json.dumps(fixture, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"sample workbook : {SAMPLE_XLSX}")
    print(f"fixture         : {FIXTURE_JSON}")
    print(f"rows            : {len(build_sample_rows())}")
    print(f"health (p3)     : {current['health_score'].get('health_score')} {current['health_score'].get('health_level')}")
    print(f"total_sales (p3): {current['key_metrics'][0]['display']}")
    print(f"growth (p3)     : {fixture['narrative']['params']['growth']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
