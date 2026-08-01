from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel

router = APIRouter(prefix="/api/workbook", tags=["workbook"])

UPLOAD_DIR = Path("storage/uploads")


# ── Shared ──────────────────────────────────────────────

class InspectRequest(BaseModel):
    saved_filename: str


# ── Inspect ─────────────────────────────────────────────

class SheetInfo(BaseModel):
    name: str
    row_count: int
    column_count: int
    merged_cell_count: int
    hidden: bool
    max_used_range: str


class InspectResponse(BaseModel):
    workbook_name: str
    worksheet_count: int
    worksheet_names: list[str]
    active_worksheet: str
    properties: dict[str, str]
    sheets: list[SheetInfo]


# ── Extract ─────────────────────────────────────────────

class DataRow(BaseModel):
    row_index: int
    values: list[Any]


class SheetData(BaseModel):
    name: str
    row_count: int
    column_count: int
    headers: list[str]
    rows: list[DataRow]


class ExtractResponse(BaseModel):
    workbook: str
    sheets: list[SheetData]


# ── Helpers ─────────────────────────────────────────────

def _col_letter(n: int) -> str:
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _validate_file(file_path: Path) -> None:
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    if file_path.suffix.lower() not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="Unsupported file type")


# ── Routes ──────────────────────────────────────────────

@router.post("/inspect", response_model=InspectResponse)
def inspect_workbook(req: InspectRequest):
    file_path = UPLOAD_DIR / req.saved_filename
    _validate_file(file_path)

    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open workbook: {e}")

    props: dict[str, str] = {}
    try:
        if wb.properties:
            props["creator"] = wb.properties.creator or ""
            props["title"] = wb.properties.title or ""
            props["created"] = str(wb.properties.created) if wb.properties.created else ""
    except Exception:
        pass

    sheets: list[SheetInfo] = []
    for ws in wb.worksheets:
        try:
            merged_count = len(ws.merged_cells.ranges)
        except Exception:
            merged_count = 0

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        try:
            sheet_hidden = ws.sheet_state == "hidden"
        except Exception:
            sheet_hidden = False

        try:
            used_range = f"A1:{_col_letter(max_col)}{max_row}" if max_row and max_col else "A1"
        except Exception:
            used_range = "A1"

        sheets.append(SheetInfo(
            name=ws.title,
            row_count=max_row,
            column_count=max_col,
            merged_cell_count=merged_count,
            hidden=sheet_hidden,
            max_used_range=used_range,
        ))

    try:
        active = wb.active.title if wb.active else ""
    except Exception:
        active = ""

    wb.close()

    return InspectResponse(
        workbook_name=file_path.name,
        worksheet_count=len(wb.worksheets),
        worksheet_names=[s.title for s in wb.worksheets],
        active_worksheet=active,
        properties=props,
        sheets=sheets,
    )


@router.post("/extract", response_model=ExtractResponse)
def extract_workbook(req: InspectRequest):
    file_path = UPLOAD_DIR / req.saved_filename
    _validate_file(file_path)

    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open workbook: {e}")

    all_sheets: list[SheetData] = []

    for ws in wb.worksheets:
        headers: list[str] = []
        rows: list[DataRow] = []
        header_found = False
        row_idx = 0

        for row in wb[ws.title].iter_rows(values_only=True):
            row_idx += 1
            values = [v if v is not None else "" for v in row]

            if all(v == "" for v in values):
                continue

            if not header_found:
                headers = [str(v) for v in values]
                header_found = True
            else:
                rows.append(DataRow(row_index=row_idx, values=values))

        all_sheets.append(SheetData(
            name=ws.title,
            row_count=len(rows),
            column_count=len(headers),
            headers=headers,
            rows=rows,
        ))

    wb.close()

    return ExtractResponse(
        workbook=file_path.name,
        sheets=all_sheets,
    )

# ── Semantic ────────────────────────────────────────────

class ColumnInfo(BaseModel):
    name: str
    type: str  # text | number | date | boolean | empty | unknown


class SemanticSheet(BaseModel):
    sheet: str
    row_count: int
    column_count: int
    columns: list[ColumnInfo]


class SemanticResponse(BaseModel):
    workbook: str
    tables: list[SemanticSheet]


def _detect_type(values: list[Any]) -> str:
    """Detect the generic type of a column from its non-empty values."""
    non_empty = [v for v in values if v not in (None, "", "None")]
    if not non_empty:
        return "empty"

    types = {type(v) for v in non_empty}

    # All same native type
    if types == {int} or types == {float} or types == {int, float}:
        return "number"
    if types == {str}:
        return "text"
    if types == {bool}:
        return "boolean"
    if all(hasattr(v, "strftime") for v in non_empty):
        return "date"

    # Try to detect dates from string patterns
    if types == {str}:
        return "text"

    return "unknown"


@router.post("/semantic", response_model=SemanticResponse)
def semantic_dataset(req: InspectRequest):
    file_path = UPLOAD_DIR / req.saved_filename
    _validate_file(file_path)

    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open workbook: {e}")

    tables: list[SemanticSheet] = []

    for ws in wb.worksheets:
        headers: list[str] = []
        columns_data: dict[int, list[Any]] = {}
        header_found = False
        data_rows = 0

        for row in wb[ws.title].iter_rows(values_only=True):
            values = [v if v is not None else "" for v in row]

            if all(v == "" for v in values):
                continue

            if not header_found:
                headers = [str(v) for v in values]
                columns_data = {i: [] for i in range(len(headers))}
                header_found = True
            else:
                data_rows += 1
                for i, v in enumerate(values):
                    if i in columns_data:
                        columns_data[i].append(v)

        # Detect column types
        columns: list[ColumnInfo] = []
        for i, name in enumerate(headers):
            col_values = columns_data.get(i, [])
            col_type = _detect_type(col_values)
            columns.append(ColumnInfo(name=name, type=col_type))

        tables.append(SemanticSheet(
            sheet=ws.title,
            row_count=data_rows,
            column_count=len(headers),
            columns=columns,
        ))

    wb.close()

    return SemanticResponse(
        workbook=file_path.name,
        tables=tables,
    )