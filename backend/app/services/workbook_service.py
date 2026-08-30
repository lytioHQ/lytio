"""Reusable workbook extraction for the async analysis runner.

Kept separate from the workbook HTTP router so background jobs can read the
canonical dataset (first sheet headers / typed columns / rows) without HTTP
coupling. The workbook router remains the public API contract.
"""

from pathlib import Path
from typing import Any

import re

from openpyxl import load_workbook

from app.core.logging_config import logger

UPLOAD_DIR = Path("storage/uploads")

_DATE_RE = re.compile(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?$")
_DATE_RE_CN = re.compile(r"^\d{4}年\d{1,2}月\d{1,2}日$")
_NUMBER_CLEAN_RE = re.compile(r"[,\s¥￥$€]|^约")


class WorkbookAccessError(RuntimeError):
    """Raised when an uploaded workbook cannot be located or parsed."""

    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


def resolve_upload_path(user_id: int, saved_filename: str) -> Path:
    return UPLOAD_DIR / str(user_id) / saved_filename


def _looks_like_date(value: Any) -> bool:
    if hasattr(value, "strftime"):
        return True
    if not isinstance(value, str):
        return False
    s = value.strip()
    return bool(_DATE_RE.match(s) or _DATE_RE_CN.match(s))


def _parse_numeric(value: Any) -> float | None:
    """Parse a numeric value, including text-formatted numbers."""
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    s = value.strip().lower()
    if not s or re.match(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", s):
        return None
    s = _NUMBER_CLEAN_RE.sub("", s)
    if s.endswith("%"):
        s = s[:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _detect_type(values: list[Any]) -> str:
    """Detect a column type, including text-formatted numbers and dates."""
    non_empty = [v for v in values if v not in (None, "", "None")]
    if not non_empty:
        return "empty"

    native_types = {type(v) for v in non_empty}
    if native_types == {bool}:
        return "boolean"
    if native_types <= {int, float}:
        return "number"
    if all(hasattr(v, "strftime") for v in non_empty):
        return "date"
    if native_types == {str}:
        numeric = sum(1 for v in non_empty if _parse_numeric(v) is not None)
        dates = sum(1 for v in non_empty if _looks_like_date(v))
        if numeric / len(non_empty) >= 0.8:
            return "number"
        if dates / len(non_empty) >= 0.8:
            return "date"
        return "text"
    return "unknown"


def _find_header_index(rows: list[list[Any]]) -> int:
    """Find the first plausible header row, skipping title rows."""
    for idx, row in enumerate(rows[:10]):
        non_empty = [v for v in row if str(v).strip()]
        if len(non_empty) < 2:
            continue
        return idx
    return 0


def extract_canonical_dataset(user_id: int, saved_filename: str) -> dict[str, Any]:
    """Return the first sheet as workbook_name/sheet_name/headers/column_types/rows."""
    file_path = resolve_upload_path(user_id, saved_filename)
    logger.info(
        "analysis_job_workbook_lookup",
        extra={
            "event": "analysis_job",
            "user_id": user_id,
            "saved_filename": saved_filename,
            "exists": file_path.exists(),
        },
    )

    if not file_path.exists():
        raise WorkbookAccessError("missing_file", "Uploaded workbook not found.")
    if file_path.suffix.lower() not in (".xlsx", ".xls"):
        raise WorkbookAccessError("unsupported_file", "Unsupported workbook type.")

    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as exc:
        raise WorkbookAccessError("unreadable_file", f"Failed to open workbook: {exc}") from exc

    try:
        if not wb.worksheets:
            raise WorkbookAccessError("empty_workbook", "Workbook has no worksheets.")

        sheet = wb.worksheets[0]
        raw_rows: list[list[Any]] = []

        for row in wb[sheet.title].iter_rows(values_only=True):
            values = [v if v is not None else "" for v in row]
            if all(v == "" for v in values):
                continue
            raw_rows.append(values)

        if not raw_rows:
            raise WorkbookAccessError("empty_workbook", "Workbook contains no header row.")

        header_index = _find_header_index(raw_rows)
        headers = [str(v) for v in raw_rows[header_index]]
        rows: list[list[Any]] = []
        columns_data: dict[int, list[Any]] = {i: [] for i in range(len(headers))}

        for row in raw_rows[header_index + 1:]:
            for i, v in enumerate(row):
                if i in columns_data:
                    columns_data[i].append(v)
            rows.append(row)

        column_types: dict[str, str] = {}
        for i, name in enumerate(headers):
            column_types[name] = _detect_type(columns_data.get(i, []))

        return {
            "workbook_name": file_path.name,
            "sheet_name": sheet.title,
            "headers": headers,
            "column_types": column_types,
            "rows": rows,
        }
    finally:
        wb.close()
